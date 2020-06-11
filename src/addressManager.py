import openpyxl

from src import generateAddress
from src import output


def generateWallet(options):

    if(options["output"] == "console"):
        counter = 1
        for i in range(int(options["num"])):
            wallet = generateAddress.generateKey()
            print(str(counter))
            output.outputConsole(wallet)
            counter += 1
        print("Success generate!\n")

    elif(options["output"] == ".txt"):
        counter = 1
        file = open("wallet/Wallet.txt", "w")
        for i in range(int(options["num"])):
            wallet = generateAddress.generateKey()
            file.write(str(counter)+"\n")
            output.outputText(wallet, file)
            counter += 1
        file.close()
        print("Success generate!\n")

    elif(options["output"] == ".xlsx"):
        counter = 1
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Eth_wallet'
        sheet.column_dimensions['A'].width = 3
        sheet.column_dimensions['B'].width = 64
        sheet.column_dimensions['C'].width = 128
        sheet.column_dimensions['D'].width = 40
        sheet.column_dimensions['E'].width = 40
        sheet.cell(row=1, column=2, value="private key")
        sheet.cell(row=1, column=3, value="public key")
        sheet.cell(row=1, column=4, value="address")
        sheet.cell(row=1, column=5, value="address(EIP-55)")
        for i in range(int(options["num"])):
            wallet = generateAddress.generateKey()
            output.outputXlsx(wallet, sheet, i+2)
        wb.save('wallet/EthWallet.xlsx')
        print("Success generate!\n")
